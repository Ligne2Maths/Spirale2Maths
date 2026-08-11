# -*- coding: utf-8 -*-
"""
remplir_durees_youtube.py

Parcourt tous les fichiers .xlsx du dossier `contenu/` et complète, sur les
deux feuilles, les cellules calculables manquantes :

1. Colonne "Code" de la feuille "Liste" (1 SF par ligne) :
   Code = Préfixe + ".SF" + n°Chapitre + "." + n°SF
   (à l'origine une formule Excel — voir la note plus bas).

2. Colonnes "Code" et "Nom" de la feuille "Plannification" (planning par
   date) : recopiées depuis la ligne correspondante de "Liste" (même numéro
   de ligne, comme le faisait la formule d'origine `=Liste!D2`). Un SF
   présent dans "Liste" mais dont la ligne est encore vide dans
   "Plannification" y est donc ajouté (code + titre), prêt à être coché.

3. Colonne "Niveau N fin" de la feuille "Liste" (durée en secondes) : pour
   chaque niveau disponible (colonnes "Niveau N YT"), si "Niveau N fin"
   (2 colonnes plus loin : YT, DG, fin) est vide et qu'un ID vidéo YouTube
   est renseigné, la durée est récupérée via yt-dlp (aucune clé API requise)
   et écrite en secondes.

Toutes les colonnes sont repérées par leur en-tête (accents/espaces ignorés),
pas par une position fixe, pour rester robuste si l'ordre des colonnes change.

Pourquoi "Code" avait disparu (rappel) :
    Ces colonnes étaient à l'origine des formules Excel. `openpyxl` ne les
    évalue jamais : à chaque enregistrement, la formule est conservée mais sa
    valeur mise en cache est perdue. Comme le site (SheetJS) et ce script ne
    lisent que des valeurs, la colonne apparaissait vide. Ce script écrit
    désormais des valeurs figées (pas des formules), donc ce problème ne peut
    plus se reproduire pour "Code" une fois complété.

yt-dlp n'expose pas d'endpoint groupé comme l'API officielle : chaque vidéo
nécessite sa propre requête. Les IDs sont dédupliqués au préalable pour
éviter d'interroger deux fois la même vidéo si elle apparaît à plusieurs
endroits (plusieurs niveaux/SF).

Dépendances :
    pip install openpyxl yt-dlp

Utilisation :
    python tools/remplir_durees_youtube.py
    python tools/remplir_durees_youtube.py --dry-run   (liste sans interroger yt-dlp ni écrire)
"""

import argparse
import re
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Le module 'openpyxl' est requis. Installez-le avec :")
    print("    pip install openpyxl")
    sys.exit(1)

try:
    import yt_dlp
except ImportError:
    print("Le module 'yt-dlp' est requis. Installez-le avec :")
    print("    pip install yt-dlp")
    sys.exit(1)


NIVEAU_YT_RE = re.compile(r"^\s*niveau\s+(\d+)\s+yt\s*$", re.IGNORECASE)
YOUTUBE_URL_ID_RE = re.compile(r"(?:/embed/|watch\?v=|youtu\.be/)([^?&/\s]+)")


def normaliser_entete(texte):
    """Normalise un intitulé de colonne pour la comparaison : minuscules,
    accents retirés, ponctuation (°, etc.) réduite à des espaces. Permet de
    repérer une colonne par son nom ("Préfixe", "n° Chapitre"...) sans être
    sensible aux accents ou à la casse."""
    t = str(texte or "").strip().lower()
    t = unicodedata.normalize("NFKD", t)
    t = "".join(c for c in t if not unicodedata.combining(c))
    t = re.sub(r"[^a-z0-9]+", " ", t).strip()
    return t


def trouver_colonne(header_row, nom_normalise):
    """Renvoie l'index 1-based de la première colonne dont l'en-tête
    normalisé correspond à `nom_normalise`, ou None si absente."""
    for idx, val in enumerate(header_row, start=1):
        if normaliser_entete(val) == nom_normalise:
            return idx
    return None


def extraire_id_youtube(valeur):
    """Renvoie l'ID vidéo à partir du contenu d'une cellule "Niveau N YT"."""
    if valeur is None:
        return None
    texte = str(valeur).strip()
    if not texte:
        return None
    if "youtube.com" in texte or "youtu.be" in texte:
        m = YOUTUBE_URL_ID_RE.search(texte)
        return m.group(1) if m else texte
    return texte


def get_feuille_liste(wb):
    """Renvoie la feuille "Liste" (ou la première feuille à défaut)."""
    for nom in wb.sheetnames:
        if nom.strip().lower() == "liste":
            return wb[nom]
    return wb[wb.sheetnames[0]]


def get_feuille_plannification(wb):
    """Renvoie la feuille "Plannification" (ou la deuxième feuille à défaut)."""
    for nom in wb.sheetnames:
        if nom.strip().lower() == "plannification":
            return wb[nom]
    return wb[wb.sheetnames[1]] if len(wb.sheetnames) > 1 else None


def header_de(ws):
    return [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]


def detecter_colonnes_niveaux(header_row):
    """Repère les colonnes "Niveau N YT" et déduit la colonne "Niveau N fin"
    associée (2 colonnes plus loin : YT, DG, fin)."""
    niveaux = []
    for idx, val in enumerate(header_row, start=1):
        if not val:
            continue
        m = NIVEAU_YT_RE.match(str(val))
        if not m:
            continue
        n = int(m.group(1))
        fin_col = idx + 2
        fin_header = header_row[fin_col - 1] if fin_col - 1 < len(header_row) else None
        if not fin_header or "fin" not in str(fin_header).lower():
            print(
                "Avertissement : colonne 'fin' attendue en position {} pour "
                "'Niveau {} YT', mais en-tête trouvé = {!r}. Vérifiez la structure "
                "du tableur.".format(fin_col, n, fin_header)
            )
        niveaux.append({"niveau": n, "yt_col": idx, "fin_col": fin_col})
    return niveaux


def ouvrir_classeurs(chemins):
    """Ouvre tous les fichiers .xlsx (une seule fois, réutilisé pour les
    codes et les durées). Renvoie un dict {chemin: classeur}."""
    classeurs = {}
    for chemin in chemins:
        classeurs[chemin] = openpyxl.load_workbook(chemin)
    return classeurs


def collecter_taches_codes(classeurs):
    """Repère, sur les 2 feuilles de chaque classeur, les cellules "Code"
    (et "Nom" côté planning) vides à compléter. Renvoie une liste de tâches
    {chemin, feuille, ligne, col, valeur, quoi}."""
    taches = []

    for chemin, wb in classeurs.items():
        ws_liste = get_feuille_liste(wb)
        header_liste = header_de(ws_liste)

        col_prefixe = trouver_colonne(header_liste, "prefixe")
        col_chapitre = trouver_colonne(header_liste, "n chapitre")
        col_sf = trouver_colonne(header_liste, "n sf")
        col_code_liste = trouver_colonne(header_liste, "code")
        col_titre_liste = trouver_colonne(header_liste, "titre")

        if not all([col_prefixe, col_chapitre, col_sf, col_code_liste]):
            print(
                "{} : colonnes Préfixe / n° Chapitre / n° SF / Code introuvables "
                "dans la feuille '{}', complétion des codes ignorée pour ce "
                "fichier.".format(chemin, ws_liste.title)
            )
            continue

        sf_par_ligne = {}  # ligne -> (code, titre) — déjà présents ou calculés
        for row_idx in range(2, ws_liste.max_row + 1):
            prefixe = ws_liste.cell(row=row_idx, column=col_prefixe).value
            n_chap = ws_liste.cell(row=row_idx, column=col_chapitre).value
            n_sf = ws_liste.cell(row=row_idx, column=col_sf).value
            if prefixe in (None, "") or n_chap in (None, "") or n_sf in (None, ""):
                continue
            try:
                code = "{}.SF{}.{}".format(str(prefixe).strip(), int(n_chap), int(n_sf))
            except (TypeError, ValueError):
                continue

            existant = ws_liste.cell(row=row_idx, column=col_code_liste).value
            if existant in (None, ""):
                taches.append(
                    {
                        "chemin": chemin,
                        "feuille": "liste",
                        "ligne": row_idx,
                        "col": col_code_liste,
                        "valeur": code,
                        "quoi": "Code",
                    }
                )
            else:
                code = str(existant).strip()

            titre = ws_liste.cell(row=row_idx, column=col_titre_liste).value if col_titre_liste else None
            sf_par_ligne[row_idx] = (code, titre)

        ws_plan = get_feuille_plannification(wb)
        if ws_plan is None:
            continue
        header_plan = header_de(ws_plan)
        col_code_plan = trouver_colonne(header_plan, "code")
        # Le titre du SF va en 2e colonne du planning ("Nom") ; on la repère par
        # son en-tête et on retombe sur la colonne 2 si l'intitulé diffère.
        col_nom_plan = trouver_colonne(header_plan, "nom") or trouver_colonne(header_plan, "titre") or 2

        if not col_code_plan:
            print(
                "{} : colonne 'Code' introuvable dans la feuille '{}', "
                "complétion ignorée pour cette feuille.".format(chemin, ws_plan.title)
            )
            continue

        # Pas de garde sur ws_plan.max_row : un SF présent dans "Liste" mais
        # absent du planning doit justement y créer sa ligne.
        for row_idx, (code, titre) in sf_par_ligne.items():
            if ws_plan.cell(row=row_idx, column=col_code_plan).value in (None, ""):
                taches.append(
                    {
                        "chemin": chemin,
                        "feuille": "plannification",
                        "ligne": row_idx,
                        "col": col_code_plan,
                        "valeur": code,
                        "quoi": "Code",
                    }
                )
            if titre not in (None, "") and ws_plan.cell(row=row_idx, column=col_nom_plan).value in (None, ""):
                taches.append(
                    {
                        "chemin": chemin,
                        "feuille": "plannification",
                        "ligne": row_idx,
                        "col": col_nom_plan,
                        "valeur": titre,
                        "quoi": "Nom",
                    }
                )

    return taches


def appliquer_codes(classeurs, taches_codes):
    """Écrit les codes et titres calculés dans les cellules repérées. Renvoie
    l'ensemble des chemins modifiés."""
    chemins_modifies = set()
    for t in taches_codes:
        wb = classeurs[t["chemin"]]
        ws = get_feuille_liste(wb) if t["feuille"] == "liste" else get_feuille_plannification(wb)
        ws.cell(row=t["ligne"], column=t["col"]).value = t["valeur"]
        chemins_modifies.add(t["chemin"])
    return chemins_modifies


def collecter_taches_durees(classeurs):
    """Repère, dans la feuille "Liste" de chaque classeur, les cases 'fin'
    vides avec un ID vidéo renseigné. Renvoie (taches, ids_uniques)."""
    taches = []
    ids_uniques = set()

    for chemin, wb in classeurs.items():
        ws = get_feuille_liste(wb)
        header = header_de(ws)
        niveaux = detecter_colonnes_niveaux(header)

        if not niveaux:
            print(
                "{} : aucune colonne 'Niveau N YT' trouvée dans la feuille "
                "'{}', complétion des durées ignorée pour ce fichier.".format(chemin, ws.title)
            )
            continue

        for row_idx in range(2, ws.max_row + 1):
            for niv in niveaux:
                yt_cell = ws.cell(row=row_idx, column=niv["yt_col"])
                video_id = extraire_id_youtube(yt_cell.value)
                if not video_id:
                    continue

                fin_cell = ws.cell(row=row_idx, column=niv["fin_col"])
                if fin_cell.value not in (None, ""):
                    continue  # déjà renseigné

                taches.append(
                    {
                        "chemin": chemin,
                        "ligne": row_idx,
                        "fin_col": niv["fin_col"],
                        "niveau": niv["niveau"],
                        "video_id": video_id,
                    }
                )
                ids_uniques.add(video_id)

    return taches, ids_uniques


def recuperer_durees(video_ids):
    """Interroge yt-dlp vidéo par vidéo (pas de mode groupé disponible) et
    renvoie un dict {video_id: duree_en_secondes}."""
    durees = {}
    ids = sorted(video_ids)

    ydl_opts = {
        "quiet": True,
        "no_warnings": True,
        "skip_download": True,
        "noplaylist": True,
        "extract_flat": False,
    }

    with yt_dlp.YoutubeDL(ydl_opts) as ydl:
        for i, video_id in enumerate(ids, start=1):
            url = "https://www.youtube.com/watch?v={}".format(video_id)
            print("  [{}/{}] {}...".format(i, len(ids), video_id))
            try:
                info = ydl.extract_info(url, download=False)
            except yt_dlp.utils.DownloadError as e:
                print("Avertissement : vidéo introuvable, privée ou supprimée -> {} ({})".format(video_id, e))
                continue

            duree = info.get("duration") if info else None
            if duree is None:
                print("Avertissement : durée introuvable pour la vidéo {}".format(video_id))
                continue
            durees[video_id] = int(duree)

    return durees


def appliquer_durees(classeurs, taches_durees, durees):
    """Écrit les durées trouvées dans les cellules 'fin' repérées. Renvoie
    (nb_rempli, chemins_modifies)."""
    chemins_modifies = set()
    remplis = 0

    for tache in taches_durees:
        secondes = durees.get(tache["video_id"])
        if secondes is None:
            continue
        wb = classeurs[tache["chemin"]]
        ws = get_feuille_liste(wb)
        ws.cell(row=tache["ligne"], column=tache["fin_col"]).value = secondes
        remplis += 1
        chemins_modifies.add(tache["chemin"])

    return remplis, chemins_modifies


def avertir_si_formules(wb, chemin):
    """openpyxl n'évalue jamais les formules : à l'enregistrement, la valeur
    mise en cache de toute cellule-formule est perdue (seule la formule est
    réécrite). Le site (SheetJS) et ce script ne lisent que des valeurs, donc
    une telle cellule redeviendrait vide après cet enregistrement. On alerte
    avant que ça n'arrive, plutôt que de casser silencieusement le site
    (c'est exactement ce qui est arrivé à la colonne "Code" une première fois)."""
    trouvees = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    trouvees.append("{}!{}".format(ws.title, cell.coordinate))
    if trouvees:
        print(
            "ATTENTION : {} contient {} cellule(s) formule ({}). "
            "openpyxl va effacer leur valeur en cache lors de l'enregistrement "
            "(seule la formule sera conservée, plus sa valeur calculée) — le "
            "site ne lit que les valeurs, pas les formules. Remplacez ces "
            "formules par leur valeur figée (Copier > Collage spécial > "
            "Valeurs) avant de relancer.".format(
                chemin, len(trouvees), ", ".join(trouvees[:10]) + ("…" if len(trouvees) > 10 else "")
            )
        )


def enregistrer(classeurs, chemins_a_enregistrer):
    for chemin in chemins_a_enregistrer:
        avertir_si_formules(classeurs[chemin], chemin)
        try:
            classeurs[chemin].save(chemin)
            print("{} : enregistré.".format(chemin))
        except PermissionError:
            print(
                "Impossible d'enregistrer {} : le fichier est probablement ouvert "
                "dans Excel. Fermez-le puis relancez le script.".format(chemin)
            )


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Complète les codes ('Code' sur les 2 feuilles) et les durées "
            "('Niveau N fin', via yt-dlp, sans clé API) manquants dans les "
            "tableurs de contenu/."
        )
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="N'interroge pas yt-dlp et n'enregistre rien : liste seulement "
        "ce qui manque (codes et durées).",
    )
    parser.add_argument(
        "--contenu-dir",
        default="contenu",
        help="Dossier contenant les fichiers .xlsx (par défaut : contenu)",
    )
    args = parser.parse_args()

    contenu_dir = Path(args.contenu_dir)
    chemins = sorted(
        p for p in contenu_dir.glob("*.xlsx") if not p.name.startswith("~$")
    )

    if not chemins:
        print("Aucun fichier .xlsx trouvé dans {}.".format(contenu_dir))
        return

    print("{} fichier(s) trouvé(s) : {}".format(len(chemins), ", ".join(p.name for p in chemins)))

    classeurs = ouvrir_classeurs(chemins)

    taches_codes = collecter_taches_codes(classeurs)
    taches_durees, ids_uniques = collecter_taches_durees(classeurs)

    if not taches_codes and not taches_durees:
        print("Rien à faire : tous les codes et toutes les durées sont déjà renseignés.")
        for wb in classeurs.values():
            wb.close()
        return

    print(
        "{} case(s) 'Code'/'Nom' à compléter, {} case(s) 'fin' à remplir pour {} "
        "vidéo(s) distincte(s).".format(len(taches_codes), len(taches_durees), len(ids_uniques))
    )

    if args.dry_run:
        print("--dry-run : aucun appel yt-dlp, aucune écriture.")
        for t in taches_codes:
            print(
                "  [{}] {} ({}) ligne {} -> {}".format(
                    t["quoi"], t["chemin"].name, t["feuille"], t["ligne"], t["valeur"]
                )
            )
        for t in taches_durees:
            print(
                "  [Durée] {} ligne {} Niveau {} -> {}".format(
                    t["chemin"].name, t["ligne"], t["niveau"], t["video_id"]
                )
            )
        for wb in classeurs.values():
            wb.close()
        return

    chemins_modifies = set()

    if taches_codes:
        chemins_modifies |= appliquer_codes(classeurs, taches_codes)
        print("{} case(s) 'Code'/'Nom' complétée(s).".format(len(taches_codes)))

    if ids_uniques:
        print("Interrogation de yt-dlp (une requête par vidéo)...")
        durees = recuperer_durees(ids_uniques)
        remplis, chemins_durees = appliquer_durees(classeurs, taches_durees, durees)
        chemins_modifies |= chemins_durees
        print("{} durée(s) renseignée(s) sur {} attendue(s).".format(remplis, len(taches_durees)))

    enregistrer(classeurs, chemins_modifies)

    for wb in classeurs.values():
        wb.close()


if __name__ == "__main__":
    main()
